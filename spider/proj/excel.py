import os

import pandas as pd
from openpyxl.reader.excel import load_workbook

excel_name = 'd06.xlsx'
excel_output_path = f"{os.path.join(os.path.abspath('..'), excel_name)}"

def to_excel(data,col):
    df = pd.DataFrame(data)
    df.style.set_properties(**{'text-align': 'center'}).to_excel(excel_output_path, header=col,index=False)
    _auto_line()

def _auto_line():
    # 设置基础列宽
    base_width = 20

    wb = load_workbook(excel_output_path)
    ws = wb['Sheet1']
    os.remove(excel_output_path)
    # 调整列宽
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # 获取列字母
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = max(base_width, max_length + 2)  # 确保宽度至少为base_width
        ws.column_dimensions[column].width = adjusted_width

    # 调整行高
    for row in ws.iter_rows():
        max_height = 0
        for cell in row:
            if cell.value:
                cell_lines = str(cell.value).count('\n') + 1
                if cell_lines > max_height:
                    max_height = cell_lines
        ws.row_dimensions[row[0].row].height = (max_height * 15)

    # 保存调整后的Excel文件
    wb.save(excel_output_path)
    print(f'excel导出完成，文件路径:{excel_output_path}')