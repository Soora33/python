import os
import re
import time
from io import BytesIO

import pandas as pd
import requests
import urllib3
from bs4 import BeautifulSoup
from openpyxl.drawing.image import Image
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Alignment, Font

full_pre = re.compile(r'w/(..)')
url_id = re.compile(r'w/(.*)')
find_url = re.compile(r'(.*)w/*')
find_tag = re.compile(r'title=".*">(.*)</a></li>')
find_img_size = re.compile(r'<dt>Size</dt><dd>(.*?)</dd>')
find_img_views = re.compile(r'<dt>Views</dt><dd>(.*?)</dd>')
find_img_like = re.compile(r'title="User Favorites">(.*)</a>')

excel_name = 'image.xlsx'
error_list = []

def get_data(base_url):
    result_list = []
    print(f'正在爬取{base_url}页面信息……')
    data = ask_url(base_url)
    if data is None:
        return result_list
    img_list = data.find_all('a', class_='preview')
    print(f'页面爬取成功，正在解析信息……')
    for item in img_list:
        img_obj = []
        # 获取图片源地址
        href = item.get('href')
        img_obj.append(href)

        # 继续打开二级页面
        id = re.findall(url_id,href)
        if len(id) <= 0:
            print('页面图片路径异常')
            continue
        time.sleep(1)
        image_info = ask_url('https://wallhaven.cc/w/' + id[0])
        if image_info is None:
            continue
        # 拿到详情大图地址
        link = image_info.find_all('img',id = 'wallpaper')
        if len(link) <= 0:
            print(image_info)
            print(link)
            img_obj.append('暂无')
        else:
            img_obj.append(link[0].get('src'))

        # 继续获取TAG
        tag_list_source = image_info.find_all('li',class_ = 'tag tag-sfw')
        tag_list = []
        for tagItem in tag_list_source:
            tag = re.findall(find_tag,str(tagItem))
            if len(tag) != 0:
                tag_list.append(tag[0])
        img_obj.append(tag_list)

        # 获取图片其余信息 大小、观看数、收藏
        image_other = image_info.find_all('dl')
        if len(image_other) <= 0:
            img_obj.append('暂无')
            img_obj.append('暂无')
            img_obj.append('暂无')
        else:
            if len(image_other[0]) != 0:
                size = re.findall(find_img_size,str(image_other[0]))
                views = re.findall(find_img_views,str(image_other[0]))
                link = re.findall(find_img_like,str(image_other[0]))
                img_obj.append(size[0].split('-')[0] if len(size)>0 else '')
                img_obj.append(views[0] if len(views)>0 else '')
                img_obj.append(link[0] if len(link)>0 else '')

        result_list.append(img_obj)
    print(f'页面信息解析完成')
    return result_list

def ask_url(base_url):
    pool = urllib3.PoolManager()
    html = pool.request('GET',base_url,headers={"User-Agent": 'Mozilla / 5.0(Windows NT 10.0; Win64; x64) AppleWebKit / 537.36(KHTML, like Gecko) Chrome / 80.0.3987.122  Safari / 537.36'})
    if html.status != 200:
        error_list.append(base_url)
        print(f'已收集错误url{base_url}')
        return None
    soup = BeautifulSoup(html.data.decode('utf-8'),'html.parser')
    return soup

def to_excel(result_list):
    col = ('图片地址','图片详情','图片标签','图片大小','图片浏览数','图片收藏数')
    df = pd.DataFrame(result_list)
    df.to_excel(excel_name,sheet_name='SHEET1', index=False,header=col)
    auto_line_size()

def auto_line_size():
    # 设置基础列宽
    base_width = 20
    base_max_width = 300
    wb = load_workbook(excel_name)
    ws = wb['SHEET1']
    os.remove(excel_name)
    url_column = 'B'
    # 获取URL列的列号
    url_column_index = ws[url_column + '1'].column

    # 新增一列用于展示图片
    ws.insert_cols(url_column_index + 1)
    new_column_letter = ws.cell(row=1,column=url_column_index + 1).column_letter
    # 添加列名
    header_cell = ws[f"{new_column_letter}1"]
    header_cell.value = "图片预览"
    header_cell.font = Font(bold=True)

    # 设置第一行的单元格水平和垂直居中
    for cell in ws[1]:
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 插入图片到新增列
    for row in ws.iter_rows(min_col=url_column_index, max_col=url_column_index, min_row=2):
        cell = row[0]
        img_url = cell.value

        if img_url:
            try:
                # 下载图片
                time.sleep(1)
                response = requests.get(img_url)
                img = Image(BytesIO(response.content))

                # 获取图片原始大小
                original_width, original_height = img.width, img.height

                # 缩小图片大小（10倍）
                img.width = min(original_width / 10,300)
                img.height = min(original_height / 10,300)

                # 在新增列的单元格中插入图片
                ws.add_image(img, f"{new_column_letter}{cell.row}")

                # 调整对应行的行高
                ws.row_dimensions[cell.row].height = img.height

            except Exception as e:
                print(f"Failed to insert image from {img_url}: {e}")

    # 调整图片列的列宽
    ws.column_dimensions[new_column_letter].width = base_max_width / 7
    # 调整列宽
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # 获取列字母
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = max(base_width, max_length + 2)  # 确保宽度至少为base_width
        if column != new_column_letter:  # 跳过图片列
            ws.column_dimensions[column].width = adjusted_width

    # 保存调整后的Excel文件
    wb.save(excel_name)
    print(f'excel导出完成，文件目录{os.getcwd()}/{excel_name}')

if __name__ == '__main__':
    print("\n")
    result_list = []
    search_name = input('输入关键词（输入1自动下载排行榜图片 非1则直接搜索并下载）')
    page = input('请输入想要查询到第几页数（1页为24个图片）')
    if search_name == '1':
        for i in range(1,int(page)+1):
            if i == 1:
                base_url = 'https://wallhaven.cc/toplist'
            else:
                base_url = f'https://wallhaven.cc/toplist?page={i}'
            result = get_data(base_url)
            result_list += result
    else:
        for i in range(1,int(page)+1):
            if i == 1:
                base_url = f'https://wallhaven.cc/search?q={search_name}&categories=110&purity=100&sorting=relevance&order=desc&ai_art_filter=1'
            else:
                base_url = f'https://wallhaven.cc/search?q={search_name}&categories=110&purity=100&sorting=relevance&order=desc&ai_art_filter=1&page={i}'
            result = get_data(base_url)
            result_list += result

    # 继续对异常的url获取到数据
    print(f'继续对错误url进行二次获取，共{len(error_list)}个')
    for url in error_list:
        result = get_data(url)
        result_list += result
    print('准备将数据导入至excel，正在下载图片……')
    to_excel(result_list)