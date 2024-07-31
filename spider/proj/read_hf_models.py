import os
import re

import pandas as pd
import urllib3
from bs4 import BeautifulSoup
from openpyxl.reader.excel import load_workbook

# 定义正则
find_title = re.compile(r'<h4 class="text-md truncate font-mono text-black dark:group-hover/repo:text-yellow-500 group-hover/repo:text-indigo-600 text-smd">(.*)</h4>')
find_update_time = re.compile(r'<time[^>]*>([^<]*)</time>')
find_other_info = re.compile(r'>\s*([\d.]+[kK]?)\s*<')
find_type = re.compile(r'svg>\s*([\D.]+)\s*<span')

# 定义爬取网站和excel文件名
url = 'https://huggingface.co/models'
excel_name = 'Models.xlsx'

# 数据载体集合
resultData = []

def run(url):
    data = get_url_data(url)
    list = data.find_all('div',class_='w-full truncate')
    print(f'页面数据获取完成，开始解析目标数据，共{len(list)}条')
    for item in list:
        data = []
        item = str(item)

        title = re.findall(find_title, item)[0]
        data.append(title)

        # 拼接链接
        data.append('https://huggingface.co/' + title)

        type = find_type.findall(item)
        if len(type) == 1:
            data.append(type[0].replace('\n','').replace('\t',''))
        else:
            data.append("暂无")

        update_time = re.findall(find_update_time, item)[0]
        data.append(update_time)

        other_info = find_other_info.findall(item)
        if len(other_info) == 1:
            if other_info[0].find('k') != -1:
                data.append(other_info[0])
                data.append('暂无')
            else:
                data.append('暂无')
                data.append(other_info[0])
        else:
            data.append(other_info[0])
            data.append(other_info[1])

        resultData.append(data)

    print(f'数据解析完成，正在导出到excel……')


def get_url_data(url):
    pool = urllib3.PoolManager()
    print(f'开始爬取页面{url}数据……')
    html = pool.request('GET',url,headers={"User-Agent": "Mozilla / 5.0(Windows NT 10.0; Win64; x64) AppleWebKit / 537.36(KHTML, like Gecko) Chrome / 80.0.3987.122  Safari / 537.36"})
    if html.status == 200:
        soup = BeautifulSoup(html.data.decode('utf-8'), 'html.parser')
        return soup
    else:
        print('解析页面出错')


def to_excel(data):
    col = ('模型名','模型对应链接','模型类型','模型最后更新时间','模型下载次数','模型收藏数')
    df = pd.DataFrame(data)
    df.to_excel(excel_name, sheet_name='Models',index=False,header=col)
    auto_line_size()

def auto_line_size():
    # 设置基础列宽
    base_width = 20

    wb = load_workbook(excel_name)
    ws = wb['Models']
    os.remove(excel_name)
    # 调整列宽
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # 获取列字母
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = max(base_width, max_length + 2)  # 确保宽度至少为base_width
        ws.column_dimensions[column].width = adjusted_width

    # 调整行高
    for row in ws.iter_rows():
        max_height = 0
        for cell in row:
            if cell.value:
                cell_lines = str(cell.value).count('\n') + 1
                if cell_lines > max_height:
                    max_height = cell_lines
        ws.row_dimensions[row[0].row].height = (max_height * 15)

    # 保存调整后的Excel文件
    wb.save(excel_name)
    print(f'excel导出完成，文件目录{os.getcwd()}/{excel_name}')

if __name__ == '__main__':
    print("/n")
# 爬取 1-5 页的数据
for i in range(1,6):
    if i == 1:
        run(url)
    else:
        run(url + f'?p={i}&sort=trending')
to_excel(resultData)
