import os

import bs4
import pandas as pd
import urllib3
from openpyxl.reader.excel import load_workbook

col = []
url_list = []
result_list = []
pool = urllib3.PoolManager()

excel_path = f'{os.getcwd()}/E04.xlsx'

def get_data(soup:str):
    if soup is None:
        return None

    # 获取所有的值
    tr_list = str(soup.find_all('tr')).split(',')
    for tr in tr_list:
        tr_html = bs4.BeautifulSoup(tr,'html.parser')
        td_list = tr_html.find_all('td')
        # 初始化对象
        data = []
        for td in td_list:
            data.append(td.getText())
        # 加入到最终集合
        if data != []:
            result_list.append(data)

    return result_list

def ask_url(base_url:str):
    html = pool.request('GET',base_url,headers={"User-Agent": 'Mozilla / 5.0(Windows NT 10.0; Win64; x64) AppleWebKit / 537.36(KHTML, like Gecko) Chrome / 80.0.3987.122  Safari / 537.36'})
    if html.status == 200:
        soup = bs4.BeautifulSoup(html.data.decode('utf-8'),'html.parser')
        return soup
    return None

def to_excel(data):
    df = pd.DataFrame(data)
    df.style.set_properties(**{'text-align': 'center'}).to_excel(excel_path, header=col,index=False)

def auto_line_size():
    # 设置基础列宽
    base_width = 20

    wb = load_workbook(excel_path)
    ws = wb['Sheet1']
    os.remove(excel_path)
    # 调整列宽
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # 获取列字母
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = max(base_width, max_length + 2)  # 确保宽度至少为base_width
        ws.column_dimensions[column].width = adjusted_width

    # 调整行高
    for row in ws.iter_rows():
        max_height = 0
        for cell in row:
            if cell.value:
                cell_lines = str(cell.value).count('\n') + 1
                if cell_lines > max_height:
                    max_height = cell_lines
        ws.row_dimensions[row[0].row].height = (max_height * 15)

    # 保存调整后的Excel文件
    wb.save(excel_path)
    print(f'excel导出完成，文件目录:{excel_path}')

if __name__ == '__main__':
    base_url = 'https://spiderbuf.cn/playground/e04/2fe6286a4e5f'
    soup = ask_url(base_url)

    # 按照th分组获取所有的字段
    tr_list = str(soup.find_all('tr')).split(',')
    tr_html = bs4.BeautifulSoup(tr_list[0],'html.parser')
    th_list = tr_html.find_all('th')
    for th in th_list:
        col.append(th.getText())

    # 判断是否存在下一页
    href_list = bs4.BeautifulSoup(str(soup.find_all('ul',class_='pagination')),'html.parser').select('a.item:not(.trap)')
    for href in href_list:
        url_list.append(f'https://spiderbuf.cn/' + str(href.get('href')))

    # 循环其他页数
    for url in url_list:
        soup = ask_url(url)
        get_data(soup)

    to_excel(result_list)
    auto_line_size()
    print(result_list)